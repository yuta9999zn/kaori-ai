"""
Smart Excel parser with multilingual semantic column detection.

Column detection uses config/language_dictionary.json as the single source
of truth for field semantics across EN / VI / JA / KO / ZH.

Scoring tiers (used in detect_columns):
  100  exact normalized match
   80  full substring (keyword ⊆ col or col ⊆ keyword)
   60  multi-word overlap (≥2 shared tokens)
   40  single-word overlap
   25  fuzzy similarity ratio > 0.85  (difflib fallback)

Confidence bands:
  HIGH    score ≥ 90
  MEDIUM  score ≥ 60
  LOW     score ≥ 30
  UNKNOWN score <  30  (user must define)
"""

import io
import json
import os
import re
import tempfile
import unicodedata
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterator

import openpyxl
import pandas as pd

from utils.wide_format import maybe_pivot, WideFormatInfo


# ---------------------------------------------------------------------------
# Language dictionary — loaded once, module-level cache
# ---------------------------------------------------------------------------

_DICT_PATH = Path(__file__).parent.parent / "config" / "language_dictionary.json"
_FLAT_DICT: dict[str, list[str]] | None = None          # {canonical: [all aliases]}
_DICT_META: dict[str, dict] | None = None               # raw field metadata


def _load_language_dict() -> dict[str, list[str]]:
    """
    Load and flatten the language dictionary.
    Returns {canonical_name: [every alias across all languages]}.
    Cached after first call.
    """
    global _FLAT_DICT, _DICT_META
    if _FLAT_DICT is not None:
        return _FLAT_DICT

    try:
        with open(_DICT_PATH, encoding="utf-8") as f:
            raw = json.load(f)
    except FileNotFoundError:
        # Graceful degradation — callers can still pass col_spec directly
        _FLAT_DICT = {}
        _DICT_META = {}
        return _FLAT_DICT

    _DICT_META = raw.get("fields", {})
    result: dict[str, list[str]] = {}
    for canonical, field_data in _DICT_META.items():
        aliases: list[str] = []
        for lang_aliases in field_data.get("aliases", {}).values():
            aliases.extend(lang_aliases)
        result[canonical] = aliases

    _FLAT_DICT = result
    return _FLAT_DICT


def _dict_description(canonical: str) -> str:
    """Human-readable description of a canonical field, for reporting."""
    if _DICT_META and canonical in _DICT_META:
        return _DICT_META[canonical].get("description", "")
    return ""


def _alias_language(canonical: str, alias: str) -> str:
    """Return the language code of a matched alias, e.g. 'ja', 'vi'."""
    if not _DICT_META or canonical not in _DICT_META:
        return "?"
    for lang, aliases in _DICT_META[canonical].get("aliases", {}).items():
        if alias in aliases:
            return lang
    return "?"


# ---------------------------------------------------------------------------
# Text normalisation
# ---------------------------------------------------------------------------

def _norm(text) -> str:
    """
    Normalise text for comparison:
    1. NFKC — converts full-width ASCII/CJK compatibility forms → standard
    2. NFD  — decomposes characters (e.g. 'à' → 'a' + combining grave)
    3. Strip combining characters (diacritics) — removes accents
    4. Lowercase, collapse whitespace

    Result: ASCII letters/digits + CJK ideographs, no accents, no full-width.
    """
    s = str(text).strip()
    s = unicodedata.normalize("NFKC", s)   # full-width → ASCII; CJK compat → standard
    s = unicodedata.normalize("NFD", s)    # decompose (enables diacritic removal)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ---------------------------------------------------------------------------
# Column match result
# ---------------------------------------------------------------------------

@dataclass
class ColumnMatch:
    raw_name: str               # actual column name from the file
    canonical: str | None       # mapped canonical field name, or None
    confidence: str             # HIGH | MEDIUM | LOW | UNKNOWN
    score: int                  # 0–100
    matched_alias: str = ""     # which alias triggered the match
    matched_lang: str = ""      # language of the matched alias
    sample_values: list[str] = field(default_factory=list)
    reason: str = ""            # human-readable explanation
    was_pivoted: bool = False   # True if column came from a wide→long pivot


# ---------------------------------------------------------------------------
# Column detection — core engine
# ---------------------------------------------------------------------------

def _score_pair(col_norm: str, kw_norm: str) -> int:
    """
    Return similarity score (0–100) for a normalised column name vs keyword.
    """
    # Tier 1: exact match
    if col_norm == kw_norm:
        return 100

    # Tier 2: full substring (either direction)
    if kw_norm and (kw_norm in col_norm or col_norm in kw_norm):
        return 80

    # Tier 3 / 4: token overlap
    col_tokens = set(col_norm.split())
    kw_tokens = set(kw_norm.split())
    # Single-char CJK tokens are too generic — require ≥2 chars
    col_tokens = {t for t in col_tokens if len(t) >= 2 or not _is_cjk(t)}
    kw_tokens  = {t for t in kw_tokens  if len(t) >= 2 or not _is_cjk(t)}

    common = col_tokens & kw_tokens
    if len(common) >= 2:
        return 60
    if len(common) == 1:
        return 40

    # Tier 5: fuzzy fallback (catches typos, minor misspellings)
    # Only apply to strings of similar length — avoids false positives
    if kw_norm and abs(len(col_norm) - len(kw_norm)) <= 4:
        ratio = SequenceMatcher(None, col_norm, kw_norm).ratio()
        if ratio > 0.85:
            return 25

    return 0


def _is_cjk(char: str) -> bool:
    """True if the first character is in CJK Unified Ideographs range."""
    if not char:
        return False
    cp = ord(char[0])
    return (0x4E00 <= cp <= 0x9FFF or   # CJK Unified
            0x3400 <= cp <= 0x4DBF or   # CJK Extension A
            0xF900 <= cp <= 0xFAFF or   # CJK Compatibility
            0xAC00 <= cp <= 0xD7A3 or   # Korean Hangul
            0x3040 <= cp <= 0x30FF)     # Hiragana + Katakana


def detect_columns(
    df: pd.DataFrame,
    col_spec: dict[str, list[str]] | None = None,
    sample_n: int = 3,
) -> list[ColumnMatch]:
    """
    Match every DataFrame column to a canonical field name.

    Sources (merged, highest priority first):
      1. col_spec passed by caller  (domain-specific hints)
      2. language_dictionary.json   (multilingual semantics)

    Returns one ColumnMatch per column. Unmatched columns get canonical=None.
    """
    # Build merged spec: caller hints take priority (prepended)
    base = _load_language_dict()
    merged: dict[str, list[str]] = {k: list(v) for k, v in base.items()}
    if col_spec:
        for canonical, keywords in col_spec.items():
            if canonical in merged:
                merged[canonical] = keywords + merged[canonical]  # caller first
            else:
                merged[canonical] = list(keywords)

    # Pre-normalise all column names once
    col_norms = {col: _norm(str(col)) for col in df.columns}

    # Score every (canonical, df_column, alias) triple
    # Store: scores[(canonical, col)] = (best_score, best_alias, best_alias_raw)
    scores: dict[tuple[str, str], tuple[int, str, str]] = {}

    for canonical, aliases in merged.items():
        for col in df.columns:
            col_norm = col_norms[col]
            best_score = 0
            best_alias_norm = ""
            best_alias_raw = ""
            for alias in aliases:
                alias_norm = _norm(alias)
                s = _score_pair(col_norm, alias_norm)
                if s > best_score:
                    best_score = s
                    best_alias_norm = alias_norm
                    best_alias_raw = alias
            if best_score > 0:
                scores[(canonical, col)] = (best_score, best_alias_norm, best_alias_raw)

    # Greedy assignment: highest-score pair first, no canonical or column reused
    assigned_canonical: dict[str, str] = {}   # canonical → col
    assigned_col: dict[str, str] = {}          # col → canonical

    sorted_pairs = sorted(
        scores.items(),
        key=lambda x: (-x[1][0], x[0][0], x[0][1])   # desc score, then alphabetical
    )
    for (canonical, col), (score, _, _) in sorted_pairs:
        if canonical in assigned_canonical or col in assigned_col:
            continue
        assigned_canonical[canonical] = col
        assigned_col[col] = canonical

    # Build result list — one entry per df column
    def _samples(col: str) -> list[str]:
        vals = df[col].dropna().astype(str)
        vals = [v for v in vals if v.strip().lower() not in ("", "nan", "none", "-")]
        return vals[:sample_n]

    def _confidence(score: int) -> str:
        if score >= 90:
            return "HIGH"
        if score >= 60:
            return "MEDIUM"
        if score >= 30:
            return "LOW"
        return "UNKNOWN"

    results: list[ColumnMatch] = []
    for col in df.columns:
        canonical = assigned_col.get(col)
        if canonical:
            score, alias_norm, alias_raw = scores[(canonical, col)]
            lang = _alias_language(canonical, alias_raw)
            reason = _build_reason(col, canonical, score, alias_raw, alias_norm, lang)
        else:
            score = 0
            alias_raw = ""
            lang = ""
            reason = "No keyword match found in any language — set user_override in mapping file"

        results.append(ColumnMatch(
            raw_name=str(col),
            canonical=canonical,
            confidence=_confidence(score) if canonical else "UNKNOWN",
            score=score,
            matched_alias=alias_raw,
            matched_lang=lang,
            sample_values=_samples(col),
            reason=reason,
        ))

    return results


def _build_reason(col: str, canonical: str, score: int,
                  alias_raw: str, alias_norm: str, lang: str) -> str:
    lang_labels = {"en": "English", "vi": "Vietnamese", "ja": "Japanese",
                   "ko": "Korean", "zh": "Chinese", "?": "unknown language"}
    lang_label = lang_labels.get(lang, lang)

    if score == 100:
        how = "exact match"
    elif score == 80:
        how = "substring match"
    elif score >= 60:
        how = "word overlap"
    elif score >= 40:
        how = "partial word match"
    else:
        how = "fuzzy match"

    if alias_raw == _norm(col):
        return f"Matched '{canonical}' — {how} on '{alias_raw}' [{lang_label}] (score {score})"
    else:
        return (f"Matched '{canonical}' — {how} on {lang_label} alias "
                f"'{alias_raw}' for column '{col}' (score {score})")


# ---------------------------------------------------------------------------
# Header row detection  (also driven by language dictionary)
# ---------------------------------------------------------------------------

def _build_header_keywords() -> set[str]:
    """Flatten all dictionary aliases into a set for header scoring."""
    flat = _load_language_dict()
    keywords: set[str] = set()
    for aliases in flat.values():
        for alias in aliases:
            keywords.add(_norm(alias))
    # Always include short single-word sentinels that are common but not in the dict
    keywords.update({"no", "stt", "id", "#", "sl", "sdt", "nv", "ck", "tm"})
    return keywords


# Built lazily — populated on first call to _header_score
_HEADER_KEYWORDS: set[str] | None = None


def _header_score(cells: list) -> int:
    global _HEADER_KEYWORDS
    if _HEADER_KEYWORDS is None:
        _HEADER_KEYWORDS = _build_header_keywords()

    score = 0
    for cell in cells:
        if cell is None or str(cell).strip() == "":
            continue
        val = str(cell).strip()
        norm_val = _norm(val)

        if norm_val in _HEADER_KEYWORDS:
            score += 5
            continue
        if any(kw in norm_val for kw in _HEADER_KEYWORDS):
            score += 3
            continue
        # Short non-numeric string → probably a label
        if len(val) <= 30 and not _looks_numeric(val) and not _looks_date(val):
            score += 1
        # Large number → data row, penalise
        if _looks_numeric(val):
            try:
                if float(re.sub(r"[,.\s]", "", val)) > 10_000:
                    score -= 3
            except ValueError:
                pass
    return score


def _looks_numeric(val: str) -> bool:
    cleaned = re.sub(r"[,.\s]", "", val.strip())
    return bool(re.fullmatch(r"-?\d+", cleaned))


def _looks_date(val: str) -> bool:
    try:
        pd.to_datetime(val, dayfirst=True)
        return True
    except Exception:
        return False


def find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    max_scan: int = 25,
) -> int:
    """
    Scan first max_scan rows; return 0-based index of the most likely header row.
    When scores are tied, take the LAST candidate (handles multi-row titles above headers).
    """
    scored: list[tuple[int, int]] = []
    max_row = min(ws.max_row or 1, max_scan)
    max_col = min(ws.max_column or 1, 30)

    for row_idx in range(max_row):
        cells = [ws.cell(row=row_idx + 1, column=c).value
                 for c in range(1, max_col + 1)]
        non_null = sum(1 for c in cells if c is not None and str(c).strip() != "")
        if non_null < 2:
            continue
        s = _header_score([str(c) for c in cells if c is not None])
        if s > 0:
            scored.append((row_idx, s))

    if not scored:
        return 0
    best = max(s for _, s in scored)
    candidates = [idx for idx, s in scored if s == best]
    return candidates[-1]


# ---------------------------------------------------------------------------
# Merged cell ratio
# ---------------------------------------------------------------------------

def merged_cell_ratio(ws: openpyxl.worksheet.worksheet.Worksheet) -> float:
    if not ws.merged_cells.ranges:
        return 0.0
    merged_area = sum(
        (rng.max_row - rng.min_row + 1) * (rng.max_col - rng.min_col + 1)
        for rng in ws.merged_cells.ranges
    )
    total = (ws.max_row or 1) * (ws.max_column or 1)
    return merged_area / max(total, 1)


# ---------------------------------------------------------------------------
# Numeric data check
# ---------------------------------------------------------------------------

def has_numeric_data(df: pd.DataFrame, min_values: int = 2) -> bool:
    for col in df.columns:
        series = df[col].dropna().astype(str)
        count = sum(
            1 for v in series
            if re.fullmatch(r"-?\d{3,}", re.sub(r"[,.\s]", "", v.strip()))
        )
        if count >= min_values:
            return True
    return False


# ---------------------------------------------------------------------------
# Sheet loader
# ---------------------------------------------------------------------------

_MERGED_SKIP_THRESHOLD = 0.35
_MIN_DATA_ROWS = 1


def load_sheet_smart(
    filepath: str,
    sheet_name: str,
    col_spec: dict[str, list[str]] | None = None,
    logger=None,
) -> tuple[pd.DataFrame, list[ColumnMatch], int, WideFormatInfo | None] | None:
    """
    Load one sheet with automatic header detection + multilingual column matching.
    Wide-format sheets (repeating column groups) are pivoted to long format first.

    Returns (df, column_matches, header_row_idx, pivot_info) or None if skipped.
    pivot_info is None for normal sheets, WideFormatInfo for pivoted sheets.
    Skipped sheets are always logged with a reason.
    """

    def _log(msg: str) -> None:
        if logger:
            logger.info(msg)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    except Exception as e:
        _log(f"  [SKIP] Cannot open '{filepath}': {e}")
        return None

    if sheet_name not in wb.sheetnames:
        _log(f"  [SKIP] Sheet '{sheet_name}' not found in workbook")
        wb.close()
        return None

    ws = wb[sheet_name]

    if not ws.max_row or ws.max_row < 2:
        _log(f"  [SKIP] Sheet '{sheet_name}': empty (max_row={ws.max_row})")
        wb.close()
        return None

    mc_ratio = merged_cell_ratio(ws)
    if mc_ratio > _MERGED_SKIP_THRESHOLD:
        _log(f"  [SKIP] Sheet '{sheet_name}': {mc_ratio:.0%} merged cells — "
             f"layout/report sheet, not a data table")
        wb.close()
        return None

    header_row_idx = find_header_row(ws)
    _log(f"  [INFO] Sheet '{sheet_name}': "
         f"header detected at row {header_row_idx + 1}  "
         f"(merged_ratio={mc_ratio:.0%})")
    wb.close()

    try:
        df = pd.read_excel(
            filepath, sheet_name=sheet_name,
            header=header_row_idx, dtype=str, engine="openpyxl",
        )
    except Exception as e:
        _log(f"  [SKIP] Sheet '{sheet_name}': pandas read failed — {e}")
        return None

    # Drop entirely empty rows and columns
    df = df.dropna(how="all").reset_index(drop=True)
    df = df.dropna(axis=1, how="all")

    def _row_empty(row) -> bool:
        return all(str(v).strip().lower() in ("", "nan", "none", "-") for v in row)
    df = df[~df.apply(_row_empty, axis=1)].reset_index(drop=True)

    if len(df) < _MIN_DATA_ROWS:
        _log(f"  [SKIP] Sheet '{sheet_name}': no data rows after header")
        return None

    if not has_numeric_data(df):
        _log(f"  [SKIP] Sheet '{sheet_name}': no numeric columns detected — "
             f"columns: {list(df.columns)[:8]}")
        return None

    # Wide-format detection and pivot (before column mapping)
    pivot_info: WideFormatInfo | None = None
    df, pivot_info = maybe_pivot(df, filepath, sheet_name, header_row_idx, logger=logger)

    matches = detect_columns(df, col_spec, sample_n=3)
    if pivot_info is not None:
        for m in matches:
            m.was_pivoted = True

    # Log summary
    matched   = [m for m in matches if m.canonical]
    unmatched = [m for m in matches if not m.canonical]
    _log(f"  [INFO] Sheet '{sheet_name}': {len(df)} rows  "
         f"matched={len(matched)} unmatched={len(unmatched)}")
    for m in matched:
        _log(f"    ✓ '{m.raw_name}' → {m.canonical} [{m.confidence}]  {m.reason}")
    for m in unmatched:
        _log(f"    ✗ '{m.raw_name}' → UNKNOWN  samples={m.sample_values[:2]}")

    return df, matches, header_row_idx, pivot_info


def iter_sheets(
    filepath: str,
    col_spec: dict[str, list[str]] | None = None,
    logger=None,
) -> Iterator[tuple[str, pd.DataFrame, list[ColumnMatch], int, WideFormatInfo | None]]:
    """Yield (sheet_name, df, column_matches, header_row_idx, pivot_info) for every usable sheet.

    Wide-format sheets are transparently pivoted before yielding — callers see
    a long-format df and matches against the metric column names.
    pivot_info is None for regular sheets, WideFormatInfo for pivoted ones.
    """
    try:
        xls = pd.ExcelFile(filepath, engine="openpyxl")
    except Exception as e:
        if logger:
            logger.error(f"Cannot open '{filepath}': {e}")
        return

    for sheet in xls.sheet_names:
        result = load_sheet_smart(filepath, sheet, col_spec, logger)
        if result is not None:
            df, matches, header_row, pivot_info = result
            yield sheet, df, matches, header_row, pivot_info


# ---------------------------------------------------------------------------
# Bronze ingestor adapter — wraps iter_sheets so services/data-pipeline can
# parse uploaded bytes (UploadFile → BytesIO) without writing the file
# anywhere visible. iter_sheets/load_sheet_smart need a real path because
# pd.ExcelFile and openpyxl.load_workbook both want one — so we stage the
# bytes in a NamedTemporaryFile, parse, and unlink.
# ---------------------------------------------------------------------------

def _detect_encoding(data: bytes) -> str:
    """Best-effort character-encoding detection for an uploaded CSV/TSV.

    latin-1 decodes ANY byte stream, so the old utf-8 → latin-1 fallback
    silently turned cp1252 / cp1250 files (£, €, Windows-Vietnamese) into
    mojibake. Strategy (more robust than a raw charset_normalizer call, which
    mis-guesses utf-16 on short ASCII-heavy buffers):

      1. Clean UTF-8 (incl. a BOM) is unambiguous → take it. utf-8-sig first so
         the BOM is stripped.
      2. Otherwise it's a legacy single-byte codepage — ask charset_normalizer,
         but ignore wide (utf-16/32) guesses; our CSVs here are single-byte.
      3. Deterministic fallback: cp1252 → cp1250 → latin-1.

    Ported from NNL-Harness profiler._detect_encoding, hardened for small files.
    """
    for enc in ("utf-8-sig", "utf-8"):
        try:
            data.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    try:
        from charset_normalizer import from_bytes
        best = from_bytes(data).best()
        if best and best.encoding and not best.encoding.replace("-", "_").startswith(
            ("utf_16", "utf_32")
        ):
            return best.encoding
    except Exception:
        pass
    for enc in ("cp1252", "cp1250", "latin-1"):
        try:
            data.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


class ExcelParser:
    """Stateless adapter used by bronze/ingestor.py.

    parse(source, filename=None) → list[dict] — one entry per usable sheet,
    shaped for the bronze_files / bronze_rows insert path:
        {sheet_name, purpose, language, header_row, rows, col_count}

    All cell values are coerced to JSON-safe primitives so downstream
    `sorted(row.items())` (row-hash) and `json.dumps(row)` (insert payload)
    don't blow up on datetime/Timestamp/NaN/bytes returned by openpyxl/pandas.
    """

    def parse(self, source, filename: str | None = None) -> list[dict]:
        ext = Path(filename).suffix.lower() if filename else ""
        data = source.read() if hasattr(source, "read") else source
        if isinstance(data, str):
            data = data.encode("utf-8")

        if ext in (".csv", ".tsv"):
            return self._parse_csv(data, ext)
        return self._parse_excel(data, ext or ".xlsx")

    @staticmethod
    def _jsonify_row(row: dict) -> dict:
        out = {}
        for k, v in row.items():
            key = str(k) if not isinstance(k, str) else k
            out[key] = _to_json_primitive(v)
        return out

    @classmethod
    def _parse_csv(cls, data: bytes, ext: str) -> list[dict]:
        sep = "\t" if ext == ".tsv" else ","
        enc = _detect_encoding(data)
        try:
            df = pd.read_csv(io.BytesIO(data), sep=sep, dtype=str, encoding=enc)
        except (UnicodeDecodeError, LookupError):
            df = pd.read_csv(io.BytesIO(data), sep=sep, dtype=str, encoding="latin-1")
        df = df.dropna(how="all").reset_index(drop=True)
        df = df.dropna(axis=1, how="all")
        rows = [cls._jsonify_row(r) for r in df.fillna("").to_dict("records")]
        return [{
            "sheet_name": "csv",
            "purpose": None,
            "language": "unknown",
            "header_row": 0,
            "rows": rows,
            "col_count": len(df.columns),
        }]

    @classmethod
    def _parse_excel(cls, data: bytes, ext: str) -> list[dict]:
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tf:
            tf.write(data)
            tmp_path = tf.name
        try:
            sheets: list[dict] = []
            for sheet_name, df, matches, header_row, _pivot in iter_sheets(tmp_path):
                lang_counter: dict[str, int] = {}
                for m in matches:
                    if m.canonical and m.matched_lang and m.matched_lang != "?":
                        lang_counter[m.matched_lang] = lang_counter.get(m.matched_lang, 0) + 1
                language = max(lang_counter, key=lang_counter.get) if lang_counter else "unknown"
                rows = [cls._jsonify_row(r) for r in df.to_dict("records")]
                sheets.append({
                    "sheet_name": sheet_name,
                    "purpose": None,
                    "language": language,
                    "header_row": header_row,
                    "rows": rows,
                    "col_count": len(df.columns),
                })
            return sheets
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def _to_json_primitive(v):
    """Coerce one cell value into something `json.dumps` and `sorted` can
    handle. Excel/pandas hands us datetime, pandas.Timestamp, NaT, NaN,
    Decimal, bytes — none of which are uniformly JSON-safe or comparable
    against str."""
    import datetime as _dt
    import math

    if v is None:
        return ""
    # pandas NaT / NaN check (NaT is also a float NaN under the hood)
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(v, (str, int, bool)):
        return v
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return ""
        return v
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, bytes):
        try:
            return v.decode("utf-8")
        except UnicodeDecodeError:
            return v.decode("latin-1", errors="replace")
    if hasattr(v, "isoformat"):  # pandas.Timestamp falls here too via datetime, but defensive
        try:
            return v.isoformat()
        except Exception:
            pass
    return str(v)

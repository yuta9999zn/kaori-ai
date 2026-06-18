"""
Excel filesystem watcher connector — PM-EVT-002.

Phase 1.5 P15-S9 D4b — replaces the P1-S3 NotImplementedError sentinel
with a one-shot polling implementation: glob the watched folder,
inspect each matching file's mtime + size + workbook metadata, emit a
NormalizedEvent per file modified within the [since, until) window.

Why one-shot polling instead of inotify/watchdog
================================================
The Process Mining session runner (Phase 1.5+ caller) drives
extraction on a schedule (Temporal cron) — it asks the connector
"give me everything modified in the last 5 minutes" and the connector
returns a finite stream. A live watchdog process complicates lifecycle
(thread vs asyncio, restart-on-crash, dropped events on the boundary)
without any acceptance-criterion need for sub-minute latency.

Workbook metadata
=================
openpyxl exposes ``creator`` (file creator) + ``lastModifiedBy`` +
``modified`` from the OOXML core properties. Phase 1.5 ships those
three; Phase 1.5+ adds column schema fingerprinting (compare against
the workflow's gold output schema) for AI-SIG-004 workaround
detection. The connector emits the metadata on the event payload so
the Process Mining consumer + AI-SIG-004 detector both consume from
the same Kafka topic.

PII boundary
============
File paths often encode tenant info (\\\\share\\\\customer-orders\\\\...).
The path is captured verbatim in ``payload['path']`` because Process
Mining downstream uses it for case grouping; the K-5 PII redactor
runs before publish to Kafka so the ai-orchestrator consumer never
sees raw paths. payload['actor'] uses ``lastModifiedBy`` (free-text
display name) which the redactor scans for VN names + email addresses.

Config keys
===========
    watch_path        absolute path to the watched directory
    file_pattern      glob pattern relative to watch_path (e.g. '**/*.xlsx');
                      the '**' enables recursive descent.
    follow_symlinks   default False — we don't follow symlinks out of
                      the watch root, prevents accidental escape from
                      a tenant-isolated mount.
    read_workbook     default True — extract creator + lastModifiedBy.
                      Set False to skip openpyxl I/O when watching a
                      large folder (mtime/size only).
"""
from __future__ import annotations

import asyncio
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, AsyncIterator, Optional

import structlog

from ...base import Connector, NormalizedEvent

log = structlog.get_logger()


_DEFAULT_FILE_PATTERN = "**/*.xlsx"


class ExcelFilesystemConnector(Connector):
    """File watcher + Excel revision metadata reader.

    Subclasses can override ``_read_workbook_metadata`` to swap openpyxl
    for a different reader (e.g. xlrd for legacy .xls when needed —
    Phase 2). The default uses openpyxl for .xlsx.
    """

    source = "excel_filesystem"

    async def extract_events(
        self,
        *,
        since: Optional[datetime] = None,
        until: Optional[datetime] = None,
    ) -> AsyncIterator[NormalizedEvent]:
        """One-shot scan of watch_path; yield one event per matching
        file modified within [since, until).

        Both bounds are advisory — when None, the bound is treated as
        unbounded (since=epoch, until=now). Caller (Process Mining
        session runner) typically supplies both: ``since`` = checkpoint
        from previous run, ``until`` = scheduler tick boundary.
        """
        watch_path = self._resolve_watch_path()
        pattern = str(self.config.get("file_pattern", _DEFAULT_FILE_PATTERN))
        follow_symlinks = bool(self.config.get("follow_symlinks", False))
        read_workbook = bool(self.config.get("read_workbook", True))

        since_dt = _coerce_aware_utc(since) if since else None
        # until=None means unbounded ceiling (include "just-modified"
        # files whose mtime equals wall-clock now). Defaulting to
        # _now_utc() races the clock against the file's mtime — a file
        # modified microseconds before the call would be missed.
        until_dt = _coerce_aware_utc(until) if until else None

        # Path.glob doesn't take a follow_symlinks flag; resolve via
        # rglob/glob and filter ourselves so the contract stays explicit
        # without pulling in `os.walk` complexity.
        candidates = list(watch_path.glob(pattern))
        log.info(
            "excel_filesystem.scan",
            tenant_id=str(self.tenant_id),
            watch_path=str(watch_path),
            pattern=pattern,
            candidates=len(candidates),
        )

        for path in candidates:
            if not path.is_file():
                continue
            if path.is_symlink() and not follow_symlinks:
                continue

            try:
                stat = path.stat()
            except OSError as exc:
                # Permission denied or file deleted between glob + stat
                # → log + skip; don't break the whole scan.
                log.warning(
                    "excel_filesystem.stat_failed",
                    path=str(path), error=str(exc),
                )
                continue

            mtime = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
            if since_dt is not None and mtime < since_dt:
                continue
            if until_dt is not None and mtime >= until_dt:
                continue

            payload: dict[str, Any] = {
                "path": str(path),
                "size_bytes": stat.st_size,
                "mtime": mtime.isoformat(),
            }
            if read_workbook:
                payload.update(self._read_workbook_metadata(path))

            event = NormalizedEvent(
                tenant_id=self.tenant_id,
                event_id=_derive_event_id(path, mtime),
                source=self.source,
                event_type="file.modified",
                occurred_at=mtime,
                actor=str(payload.get("last_modified_by") or "")  or None,
                # Group by file path so all revisions of the same Excel
                # cluster into one Process Mining case.
                case_id=str(path),
                payload=payload,
            )
            # Yield with an explicit await so a long scan doesn't starve
            # the event loop on a thousand-file directory.
            await asyncio.sleep(0)
            yield event

    # ------------------------------------------------------------------
    # Helpers — module-private but methods so subclasses can override.
    # ------------------------------------------------------------------

    def _resolve_watch_path(self) -> Path:
        """Resolve config['watch_path'] to a Path + sanity-check it.

        The connector deliberately does NOT auto-create the directory:
        a missing watch root signals a misconfiguration (wrong tenant
        share path, mount not ready) and silently scanning an empty
        auto-created folder would mask the bug for hours.
        """
        raw = self.config.get("watch_path")
        if not raw:
            raise ValueError(
                "excel_filesystem connector requires config['watch_path'] "
                "(absolute path to the watched directory)"
            )
        path = Path(raw)
        if not path.exists():
            raise FileNotFoundError(
                f"excel_filesystem watch_path not found: {path}. "
                "Either the mount is not ready or the config points at "
                "the wrong location — fix the config rather than create "
                "the directory."
            )
        if not path.is_dir():
            raise NotADirectoryError(
                f"excel_filesystem watch_path must be a directory: {path}"
            )
        return path

    def _read_workbook_metadata(self, path: Path) -> dict[str, Any]:
        """Extract OOXML core-properties from the .xlsx file.

        openpyxl is heavy on big spreadsheets — we open with
        ``read_only=True`` + ``data_only=True`` so it skips formulas +
        styles. Failures (corrupt file, encrypted file, .xls) log and
        return ``{}`` rather than raising; an unreadable file shouldn't
        break the rest of the scan.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover - openpyxl is in requirements
            log.warning("excel_filesystem.openpyxl_missing")
            return {}

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            log.warning(
                "excel_filesystem.workbook_unreadable",
                path=str(path), error=str(exc),
            )
            return {}

        try:
            props = wb.properties
            return {
                "creator": props.creator or "",
                "last_modified_by": props.lastModifiedBy or "",
                "workbook_modified": props.modified.isoformat() if props.modified else "",
                "title": props.title or "",
                "sheet_names": list(wb.sheetnames),
            }
        finally:
            wb.close()


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------


def _derive_event_id(path: Path, mtime: datetime) -> str:
    """Stable event_id per (path, mtime).

    Same revision of the same file = same id, so a re-scan after a
    crash deduplicates correctly downstream. Format mirrors
    ``ingestion.normalizer.derive_event_id`` (source-prefixed, hex
    hash) for log greppability across the connector set.
    """
    raw = f"{path}|{mtime.isoformat()}".encode("utf-8")
    digest = hashlib.sha256(raw).hexdigest()[:16]
    return f"excel_filesystem:{digest}"


def _now_utc() -> datetime:
    return datetime.now(tz=timezone.utc)


def _coerce_aware_utc(dt: datetime) -> datetime:
    """Normalise to UTC-aware datetime so since/until comparisons don't
    silently mix naive + aware timestamps."""
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)
